"""Phase 3.8 investigation dumps — DCS Excel + MOHA HTML for focus DS areas."""
from __future__ import annotations

import html
import re
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore

ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "data" / "source" / "moha-life" / "reports"
DCS_XLSX = ROOT / "data" / "source" / "dcs-gndlist-final-2024-03-19.xlsx"


def parse_report(path: Path) -> list[list[str]]:
    text = path.read_text(encoding="utf-8", errors="replace")
    rows: list[list[str]] = []
    for m in re.finditer(r"<tr[^>]*>(.*?)</tr>", text, re.I | re.S):
        cells = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", m.group(1), re.I | re.S)
        cells = [re.sub(r"<[^>]+>", "", c) for c in cells]
        cells = [html.unescape(c).strip() for c in cells]
        if len(cells) >= 9 and re.match(r"\d+-\d+-\d+-\d+", cells[0]):
            rows.append(cells)
    return rows


def life_ds(life: str) -> str:
    parts = life.split("-")
    return "".join(parts[:3])


def dump_moha() -> None:
    files = {
        "colombo": REPORTS / "p63-d42.html",
        "nuwara": REPORTS / "p64-d47.html",
        "ampara": REPORTS / "p67-d57.html",
        "galle": REPORTS / "p65-d48.html",
        "matale": REPORTS / "p64-d46.html",
        "ratnapura": REPORTS / "p71-d65.html",
    }

    print("=== COLOMBO: Ratmalana / 1131 / 1139 ===")
    rows = parse_report(files["colombo"])
    rat = [
        r
        for r in rows
        if "Ratmalana" in r[8] or r[0].startswith("1-1-31") or r[0].startswith("1-1-39")
    ]
    print(f"rows: {len(rat)}")
    for r in rat:
        print(f"  {r[0]} | gn={r[1]} | en={r[4]} | si={r[2]} | ta={r[3]} | ds={r[8]}")

    print("\n=== AMPARA: Kalmunai / Sainthamaruthu / 5221 / 5224 / 5225 ===")
    rows = parse_report(files["ampara"])
    focus = [
        r
        for r in rows
        if r[0].startswith("5-2-21")
        or r[0].startswith("5-2-24")
        or r[0].startswith("5-2-25")
        or "Kalmunai" in r[8]
        or "Saintha" in r[8]
        or "Sainda" in r[8]
    ]
    by_ds: dict[str, list] = defaultdict(list)
    for r in focus:
        by_ds[life_ds(r[0])].append(r)
    for ds, rs in sorted(by_ds.items()):
        labels = sorted({x[8] for x in rs})
        print(f"DS {ds}: n={len(rs)}")
        for lab in labels:
            print(f"  label: {lab}")
        for r in rs:
            print(f"  {r[0]} | en={r[4]} | si={r[2]} | ta={r[3]}")

    print("\n=== NUWARA ELIYA: Kothmale / Norwood ===")
    rows = parse_report(files["nuwara"])
    focus = [
        r
        for r in rows
        if any(x in r[8] for x in ("Kothmale", "Kotmale", "Norwood"))
        or r[0].startswith(("2-3-02", "2-3-04", "2-3-14", "2-3-16", "2-3-01", "2-3-03"))
    ]
    by_ds = defaultdict(list)
    for r in focus:
        by_ds[life_ds(r[0])].append(r)
    for ds, rs in sorted(by_ds.items()):
        labels = sorted({x[8] for x in rs})
        comps = sorted({x[1].zfill(3) for x in rs})
        ens = sorted({x[4] for x in rs})
        print(f"DS {ds}: n={len(rs)} labels={labels}")
        print(f"  comps={comps}")
        print(f"  ens={ens}")


def dump_dcs() -> None:
    if load_workbook is None:
        print("openpyxl missing; skip DCS dump")
        return
    if not DCS_XLSX.exists():
        print(f"DCS missing: {DCS_XLSX}")
        return
    wb = load_workbook(DCS_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = None
    focus_ds = {
        "1131",
        "2302",
        "2314",
        "5221",
        "5225",
        "2307",
        "2310",
        "2313",
        "3136",
        "3127",
        "9118",
        "2209",
    }
    by_ds: dict[str, list] = defaultdict(list)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c is not None else "" for c in row]
            continue
        if not row or row[1] is None:
            continue
        # GND_UID, DSD_ Code columns — find by header
        data = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
        uid = str(data.get("GND_UID") or "").strip()
        if len(uid) < 7:
            continue
        # DSD code column may be 'DSD_ Code'
        dsd = str(data.get("DSD_ Code") or data.get("DSD_Code") or "").strip()
        if not dsd:
            # derive from UID: province(1)+district(2)+ds(2) but UID is 7 digits
            dsd = uid[:4]
        # normalize 2-digit DSD under district
        district = str(data.get("District_Code") or uid[1:3]).strip()
        dsd_full = district + dsd.zfill(2) if len(dsd) <= 2 else dsd
        if dsd_full not in focus_ds and uid[:4] not in focus_ds:
            continue
        key = uid[:4]
        by_ds[key].append(
            (
                uid,
                str(data.get("GND_Name") or ""),
                str(data.get("DSD_Name") or ""),
                str(data.get("GND_ Code") or data.get("GND_Code") or uid[-3:]),
            )
        )
    print("\n=== DCS focus DS ===")
    for ds, items in sorted(by_ds.items()):
        print(f"DCS DS {ds}: n={len(items)} name={items[0][2] if items else ''}")
        comps = sorted({x[3].zfill(3) for x in items})
        ens = sorted({x[1] for x in items})
        print(f"  comps={comps}")
        print(f"  ens={ens}")


if __name__ == "__main__":
    import sys

    out = ROOT / ".research" / "phase38_dump_out.txt"
    with out.open("w", encoding="utf-8") as fh:
        sys.stdout = fh  # type: ignore
        dump_moha()
        dump_dcs()
    sys.stdout = sys.__stdout__
    print(f"Wrote {out}")
